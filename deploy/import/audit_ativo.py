"""Compara planilha vs regra correta de ativo/inativo e gera relatório + SQL de correção."""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
MESES_NUM = {m: i + 1 for i, m in enumerate(MESES)}
COLS = "IJKLMNOPQRST"
ANO = 2026
XLSX = Path(r"C:\Users\Lucas Cavalcante\Downloads\FINANCEIRO SOMELI.xlsx")


def parse_honorario(h, g) -> Decimal:
    raw = h if h not in (None, "") else g
    if raw in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(raw).replace(",", "."))
    except Exception:
        return Decimal("0")


def is_number(v) -> bool:
    try:
        float(str(v).replace(",", "."))
        return True
    except Exception:
        return False


def map_tipo(flag) -> str:
    f = (flag or "S").strip().upper()
    if f == "P":
        return "pessoa_fisica"
    if f == "K":
        return "terceiros"
    return "pessoa_juridica"


def detectar_inativo(pays: dict, honorario: Decimal) -> tuple[bool, int | None, bool]:
    """
    Retorna (ativo, mes_saida 1-12 ou None, saiu_em_dia).
    Flag S/P/K NÃO define ativo — só tipo de pessoa.
    Inativo quando a planilha encerrou o cliente (traços/residual após último OK,
    ou honorário 0 sem OK e só marcas de encerramento).
    """
    last_ok_i = None
    for i, m in enumerate(MESES):
        if str(pays.get(m, "")).upper() == "OK":
            last_ok_i = i

    # Nunca teve OK
    if last_ok_i is None:
        if honorario <= 0:
            # só traços / números / vazio = fora da carteira
            return False, None, True
        # honorário > 0 sem pagamentos ainda = cliente novo/ativo
        return True, None, True

    after = MESES[last_ok_i + 1 :]
    closed = []
    for m in after:
        v = pays.get(m)
        if v is None:
            continue
        vs = str(v).strip()
        if vs == "-" or (is_number(vs) and vs.upper() != "OK"):
            closed.append(m)

    # Encerrado na planilha: meses seguintes marcados com - ou residual
    if closed and honorario <= 0:
        mes_saida = MESES_NUM[MESES[last_ok_i]]  # último mês pago; saiu após ele
        # pendências = closed months that are numeric residuals (not just '-')
        tem_debito = any(is_number(str(pays[m])) and str(pays[m]).strip() != "-" for m in closed)
        # também débito se '-' mas havia honorário histórico? com hon=0 residual numbers matter
        return False, mes_saida, not tem_debito

    if len(closed) >= 3:
        # vários meses fechados mesmo com honorário — saiu
        mes_saida = MESES_NUM[MESES[last_ok_i]]
        tem_debito = any(
            is_number(str(pays[m])) and not str(pays[m]).strip().startswith("-")
            for m in closed
            if pays.get(m) is not None and str(pays[m]).strip() != "-"
        )
        # negativos = débito
        tem_debito = tem_debito or any(
            is_number(str(pays[m])) and float(str(pays[m]).replace(",", ".")) != 0
            for m in closed
            if pays.get(m) is not None and str(pays[m]).strip() != "-"
        )
        return False, mes_saida, not tem_debito

    # Traços até o fim do ano após último OK (mesmo com honorário)
    trailing = 0
    for m in reversed(MESES):
        v = pays.get(m)
        if v is None:
            continue
        if str(v).strip() == "-":
            trailing += 1
        else:
            break
    if trailing >= 2 and last_ok_i is not None and last_ok_i <= (11 - trailing):
        mes_saida = MESES_NUM[MESES[last_ok_i]]
        return False, mes_saida, True

    return True, None, True


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["RECEITAS 2026"]
    clients = []
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        nome = str(nome).strip()
        if nome.upper() in {"TOTAIS", "TOTAL", ""}:
            continue
        flag = ws.cell(r, 5).value
        pays = {}
        for m, col in zip(MESES, COLS):
            v = ws.cell(r, openpyxl.utils.column_index_from_string(col)).value
            if v is not None:
                pays[m] = str(v).strip()
        hon = parse_honorario(ws.cell(r, 8).value, ws.cell(r, 7).value)
        # regra antiga (bugada)
        old_ativo = not ((str(flag or "").strip().upper() == "P") or (hon == 0 and sum(1 for v in pays.values() if str(v).upper() == "OK") == 0))
        # wait recreate old exactly:
        f = (flag or "").strip().upper()
        ok = sum(1 for v in pays.values() if str(v).upper() == "OK")
        if f == "P":
            old_ativo = False
        elif hon == 0 and ok == 0:
            old_ativo = False
        else:
            old_ativo = True

        ativo, mes_saida, em_dia = detectar_inativo(pays, hon)
        clients.append(
            {
                "nome": nome,
                "flag": flag,
                "tipo": map_tipo(flag),
                "hon": hon,
                "old_ativo": old_ativo,
                "ativo": ativo,
                "mes_saida": mes_saida,
                "saiu_em_dia": em_dia,
                "pays": pays,
            }
        )

    print("=== DIVERGÊNCIAS (regra antiga vs nova) ===")
    for c in clients:
        if c["old_ativo"] != c["ativo"]:
            print(
                f"{c['nome'][:40]:40} antiga={'ATIVO' if c['old_ativo'] else 'INATIVO':7} "
                f"nova={'ATIVO' if c['ativo'] else 'INATIVO':7} tipo={c['tipo']:16} "
                f"saida={c['mes_saida']} em_dia={c['saiu_em_dia']}"
            )

    print("\n=== INATIVOS CORRETOS ===")
    for c in clients:
        if not c["ativo"]:
            print(
                f"{c['nome'][:40]:40} hon={c['hon']:8} saida_mes={c['mes_saida']} "
                f"em_dia={c['saiu_em_dia']} pays={c['pays']}"
            )

    print(f"\nTotal={len(clients)} ativos={sum(1 for c in clients if c['ativo'])} "
          f"inativos={sum(1 for c in clients if not c['ativo'])}")

    # SQL correção
    lines = ["BEGIN;"]
    for c in clients:
        nome = c["nome"].replace("'", "''")
        ativo_sql = "TRUE" if c["ativo"] else "FALSE"
        tipo = c["tipo"]
        if c["mes_saida"]:
            fim = f"'{ANO}-{c['mes_saida']:02d}-01'"
            # data_fim = primeiro dia do mês SEGUINTE ao último OK? User asked from which month inactive.
            # Use first day of month AFTER last paid month as fim, or last paid month end.
            sm = c["mes_saida"] + 1
            sy = ANO
            if sm > 12:
                sm = 1
                sy += 1
            fim = f"'{sy}-{sm:02d}-01'"
        else:
            fim = "NULL" if c["ativo"] else "NULL"
        lines.append(
            f"UPDATE cliente SET ativo = {ativo_sql}, tipo_pagamento = '{tipo}', "
            f"data_fim_cobranca = {fim} "
            f"WHERE lower(razao_social) = lower('{nome}') OR lower(nome_fantasia) = lower('{nome}');"
        )
    lines.append("COMMIT;")
    out = Path(r"C:\Workspace\someli-contabilidade\deploy\import\fix_ativo_from_planilha.sql")
    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nSQL gerado: {out}")


if __name__ == "__main__":
    main()
