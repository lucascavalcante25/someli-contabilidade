"""Sincroniza pagamentos/inícios de cobrança e despesas a partir da planilha."""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
import unicodedata

import openpyxl

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
MESES_NUM = {m: i + 1 for i, m in enumerate(MESES)}
COLS = "IJKLMNOPQRST"
ANO = 2026
XLSX = Path(r"C:\Users\Lucas Cavalcante\Downloads\FINANCEIRO SOMELI.xlsx")
OUT = Path(__file__).with_name("sync_planilha_pagamentos.sql")

# Snapshot ids (mesmo do gen_fix_ativo_sql)
DB_CLIENTES = [
    (7, "ADRIELEE"),
    (8, "ALEXSANDRO RODRIGUES DO NASCIMENTO"),
    (10, "ALIANCA CONSTRUCAO LTDA"),
    (11, "AUTO PEÇAS NOVA ERA"),
    (6, "Alleanza Odontologia"),
    (9, "AÇAI TIMBU"),
    (13, "BLESSED GRÃOS"),
    (15, "CA SOLUTIONS LTDA"),
    (40, "CHURRASCARIA LOURO GRILL LTDA"),
    (16, "CONSFEL CONSTRUÇÕES LTDA"),
    (14, "Clinica miranda"),
    (18, "DANIEL - MERCADINHO COMPRE SEMPRE"),
    (17, "DJ PEDRO HENRICO LTDA"),
    (19, "DS SOLUCOES LTDA"),
    (22, "EUDINHO"),
    (23, "EVYLA ADVOGADA"),
    (24, "EXTRAPACK DISTRIBUIDORA LTDA. - ME"),
    (20, "Estrela Ester Belém"),
    (21, "Estrela Ester Curitiba"),
    (54, "F F - ACADEMIA"),
    (26, "FLANK BURGUER"),
    (25, "Facundo Motos"),
    (27, "Fra Carnes"),
    (55, "Galpão Bar"),
    (30, "IRISNEI LTDA"),
    (28, "JEFFERSON LOPES"),
    (33, "JF TECNOLOGIA LTDA"),
    (34, "JG ALMEIDA"),
    (32, "KATHARINE"),
    (29, "Kelly MEI"),
    (31, "LAESTONCAR"),
    (35, "LEBENISTE"),
    (41, "LUCAS VIEIRA"),
    (36, "Leudo"),
    (37, "Limas pastelaria"),
    (38, "MATHEUS"),
    (39, "MORALES MEDICAL SUPPORT LTDA"),
    (12, "OCTOOBRAND MKT"),
    (44, "PHOTONDEPLOY"),
    (42, "Pai e Filhos auto peças"),
    (45, "Ponto do Pratinho"),
    (43, "Preço Bom Mercadinho"),
    (51, "RESTAURANTE O GORDIM"),
    (52, "RR Albuquerque"),
    (53, "RUMOS MOVEIS"),
    (46, "Registre Clicks"),
    (47, "SALINAS"),
    (48, "Samara Psicologa"),
    (49, "Samuel Alcantara"),
    (50, "Tainara"),
]


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split())


def sql_lit(v) -> str:
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def is_number(v) -> bool:
    try:
        float(str(v).replace(",", "."))
        return True
    except Exception:
        return False


def main():
    hoje = date.today()
    mes_atual = hoje.month if hoje.year == ANO else (12 if hoje.year > ANO else 1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["RECEITAS 2026"]
    by_norm = {}
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        nome = str(nome).strip()
        if nome.upper() in {"TOTAIS", "TOTAL", ""}:
            continue
        pays = {}
        for m, col in zip(MESES, COLS):
            v = ws.cell(r, openpyxl.utils.column_index_from_string(col)).value
            if v is not None:
                pays[m] = str(v).strip()
        by_norm[norm(nome)] = pays

    lines = [
        "BEGIN;",
        "ALTER TABLE pagamento_mensal ADD COLUMN IF NOT EXISTS cobravel BOOLEAN NOT NULL DEFAULT TRUE;",
        "ALTER TABLE despesa_mensal ADD COLUMN IF NOT EXISTS valor NUMERIC(12,2);",
    ]

    for cid, db_nome in DB_CLIENTES:
        pays = by_norm.get(norm(db_nome))
        if pays is None:
            print("SEM MATCH", db_nome)
            continue

        first_bill = None
        for m in MESES:
            v = pays.get(m)
            if v is None:
                continue
            if v == "-":
                continue
            if v.upper() == "OK" or is_number(v):
                first_bill = MESES_NUM[m]
                break

        if first_bill:
            lines.append(
                f"UPDATE cliente SET data_inicio_cobranca = '{ANO}-{first_bill:02d}-01' WHERE id = {cid}; "
                f"-- {db_nome}"
            )
        else:
            lines.append(f"UPDATE cliente SET data_inicio_cobranca = NULL WHERE id = {cid}; -- {db_nome} sem meses")

        # reconcilia pagamentos do ano
        for m in MESES:
            mes = MESES_NUM[m]
            v = pays.get(m)
            if v is None:
                # remove marcação cobravel=false se existir; deixa pendente sem row
                lines.append(
                    f"DELETE FROM pagamento_mensal WHERE cliente_id = {cid} AND ano = {ANO} AND mes = {mes} "
                    f"AND cobravel = FALSE;"
                )
                continue
            if v == "-":
                lines.append(
                    f"INSERT INTO pagamento_mensal (cliente_id, mes, ano, pago, cobravel) VALUES "
                    f"({cid}, {mes}, {ANO}, FALSE, FALSE) "
                    f"ON CONFLICT (cliente_id, mes, ano) DO UPDATE SET pago = FALSE, cobravel = FALSE;"
                )
            elif v.upper() == "OK":
                lines.append(
                    f"INSERT INTO pagamento_mensal (cliente_id, mes, ano, pago, cobravel) VALUES "
                    f"({cid}, {mes}, {ANO}, TRUE, TRUE) "
                    f"ON CONFLICT (cliente_id, mes, ano) DO UPDATE SET pago = TRUE, cobravel = TRUE;"
                )
            else:
                # valor numérico positivo = dívida; zero/negativo = sem cobrança efetiva
                try:
                    num = Decimal(str(v).replace(",", "."))
                except Exception:
                    num = Decimal(0)
                if num > 0:
                    lines.append(
                        f"DELETE FROM pagamento_mensal WHERE cliente_id = {cid} AND ano = {ANO} AND mes = {mes};"
                    )
                else:
                    lines.append(
                        f"INSERT INTO pagamento_mensal (cliente_id, mes, ano, pago, cobravel) VALUES "
                        f"({cid}, {mes}, {ANO}, FALSE, FALSE) "
                        f"ON CONFLICT (cliente_id, mes, ano) DO UPDATE SET pago = FALSE, cobravel = FALSE;"
                    )

        pending = []
        for m in MESES:
            v = pays.get(m)
            if v and v != "-" and v.upper() != "OK" and is_number(v):
                try:
                    if Decimal(str(v).replace(",", ".")) > 0:
                        pending.append(f"{m}={v}")
                except Exception:
                    pass
            elif v is None and first_bill and MESES_NUM[m] >= first_bill and MESES_NUM[m] <= mes_atual:
                pending.append(f"{m}=vazio")
        if pending:
            print(f"{db_nome}: pendências planilha -> {', '.join(pending)}")

    # ---- Despesas ----
    ws = wb["DESPESAS 2026"]
    mes_map = {
        "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3, "ABRIL": 4,
        "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9,
        "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
    }
    month_headers = []
    for c in range(4, ws.max_column + 1):
        h = ws.cell(3, c).value
        if not h:
            continue
        h_up = str(h).strip().upper().replace(" ", "")
        for k, m in mes_map.items():
            if k in h_up:
                month_headers.append((c, m))
                break

    lines.append("-- Despesas: valores mensais reais; meses futuros não ficam como pagos")
    lines.append(f"UPDATE despesa_mensal SET paga = FALSE WHERE ano = {ANO} AND mes > {mes_atual};")

    for r in range(4, ws.max_row + 1):
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        nome = str(nome).strip()
        if nome.upper() in {"TOTAIS", "TOTAL"}:
            continue
        valores = {}
        for col, mes in month_headers:
            v = ws.cell(r, col).value
            if v is None:
                continue
            try:
                val = Decimal(str(v).replace(",", "."))
            except Exception:
                continue
            if val > 0:
                valores[mes] = val
        if not valores:
            continue

        # valor_mensal = último mês <= atual (ou último disponível)
        ref = max((m for m in valores if m <= mes_atual), default=max(valores))
        lines.append(
            f"UPDATE despesa SET valor_mensal = {valores[ref]} WHERE lower(descricao) = lower({sql_lit(nome)});"
        )
        for mes, val in valores.items():
            paga = "TRUE" if mes <= mes_atual else "FALSE"
            lines.append(
                f"INSERT INTO despesa_mensal (despesa_id, mes, ano, paga, valor) "
                f"SELECT id, {mes}, {ANO}, {paga}, {val} FROM despesa WHERE lower(descricao) = lower({sql_lit(nome)}) "
                f"ON CONFLICT (despesa_id, mes, ano) DO UPDATE SET paga = EXCLUDED.paga, valor = EXCLUDED.valor;"
            )

    lines.append("COMMIT;")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print("SQL:", OUT)


if __name__ == "__main__":
    main()
