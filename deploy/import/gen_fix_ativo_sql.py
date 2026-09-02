"""Gera SQL de correção ativo/data_fim a partir da planilha, usando IDs do banco."""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
MESES_NUM = {m: i + 1 for i, m in enumerate(MESES)}
COLS = "IJKLMNOPQRST"
ANO = 2026
XLSX = Path(r"C:\Users\Lucas Cavalcante\Downloads\FINANCEIRO SOMELI.xlsx")

# Snapshot do banco (id, nome) — atualizar se necessário
DB_CLIENTES = [
    (7, "ADRIELEE"),
    (8, "ALEXSANDRO RODRIGUES DO NASCIMENTO"),
    (10, "ALIANCA CONSTRUCAO LTDA"),
    (11, "AUTO PEÇAS NOVA ERA"),
    (6, "Alleanza Odontologia"),
    (9, "AÇAI TIMBU"),
    (13, "BLESSED GRÃOS"),
    (15, "CA SOLUTIONS LTDA"),
    (40, "CHURRASCARIA LOURO GRILL LTDA"),
    (16, "CONSFEL CONSTRUÇÕES LTDA"),
    (14, "Clinica miranda"),
    (18, "DANIEL - MERCADINHO COMPRE SEMPRE"),
    (17, "DJ PEDRO HENRICO LTDA"),
    (19, "DS SOLUCOES LTDA"),
    (22, "EUDINHO"),
    (23, "EVYLA ADVOGADA"),
    (24, "EXTRAPACK DISTRIBUIDORA LTDA. - ME"),
    (20, "Estrela Ester Belém"),
    (21, "Estrela Ester Curitiba"),
    (54, "F F - ACADEMIA"),
    (26, "FLANK BURGUER"),
    (25, "Facundo Motos"),
    (27, "Fra Carnes"),
    (55, "Galpão Bar"),
    (30, "IRISNEI LTDA"),
    (28, "JEFFERSON LOPES"),
    (33, "JF TECNOLOGIA LTDA"),
    (34, "JG ALMEIDA"),
    (32, "KATHARINE"),
    (29, "Kelly MEI"),
    (31, "LAESTONCAR"),
    (35, "LEBENISTE"),
    (41, "LUCAS VIEIRA"),
    (36, "Leudo"),
    (37, "Limas pastelaria"),
    (38, "MATHEUS"),
    (39, "MORALES MEDICAL SUPPORT LTDA"),
    (12, "OCTOOBRAND MKT"),
    (44, "PHOTONDEPLOY"),
    (42, "Pai e Filhos auto peças"),
    (45, "Ponto do Pratinho"),
    (43, "Preço Bom Mercadinho"),
    (51, "RESTAURANTE O GORDIM"),
    (52, "RR Albuquerque"),
    (53, "RUMOS MOVEIS"),
    (46, "Registre Clicks"),
    (47, "SALINAS"),
    (48, "Samara Psicologa"),
    (49, "Samuel Alcantara"),
    (50, "Tainara"),
]


def norm(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return "".join(ch.lower() for ch in s if ch.isalnum())


def parse_honorario(h, g) -> Decimal:
    raw = h if h not in (None, "") else g
    if raw in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(raw).replace(",", "."))
    except Exception:
        return Decimal("0")


def is_number(v) -> bool:
    try:
        float(str(v).replace(",", "."))
        return True
    except Exception:
        return False


def map_tipo(flag) -> str:
    f = (flag or "S").strip().upper()
    if f == "P":
        return "pessoa_fisica"
    if f == "K":
        return "terceiros"
    return "pessoa_juridica"


def detectar(pays: dict, honorario: Decimal):
    last_ok_i = None
    last_activity_i = None
    for i, m in enumerate(MESES):
        v = pays.get(m)
        if v is None:
            continue
        vs = str(v).strip()
        if vs == "-":
            continue  # traço = mês encerrado, não estende cobrança
        last_activity_i = i
        if vs.upper() == "OK":
            last_ok_i = i

    if last_ok_i is None and last_activity_i is None:
        return honorario > 0, None

    if last_ok_i is None:
        # só residual/traços — inativo sem mês de saída claro
        return False, None

    after = MESES[last_ok_i + 1 :]
    closed = 0
    for m in after:
        v = pays.get(m)
        if v is None:
            continue
        vs = str(v).strip()
        if vs == "-" or (is_number(vs) and vs.upper() != "OK"):
            closed += 1

    # data_fim = mês seguinte à última atividade (OK ou residual)
    fim = None
    ref = last_activity_i if last_activity_i is not None else last_ok_i
    nxt = ref + 1
    if closed >= 1 or (last_activity_i is not None and last_activity_i > last_ok_i):
        if nxt >= 12:
            fim = f"{ANO + 1}-01-01"
        else:
            fim = f"{ANO}-{MESES_NUM[MESES[nxt]]:02d}-01"

    if closed >= 1 and honorario <= 0:
        return False, fim
    if closed >= 3:
        return False, fim
    if last_activity_i is not None and last_activity_i > last_ok_i and honorario <= 0:
        return False, fim

    trailing = 0
    for m in reversed(MESES):
        v = pays.get(m)
        if v is None:
            continue
        if str(v).strip() == "-":
            trailing += 1
        else:
            break
    if trailing >= 2 and last_ok_i <= (11 - trailing):
        if fim is None:
            nxt2 = last_ok_i + 1
            fim = f"{ANO + 1}-01-01" if nxt2 >= 12 else f"{ANO}-{MESES_NUM[MESES[nxt2]]:02d}-01"
        return False, fim

    return True, None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["RECEITAS 2026"]
    by_norm = {}
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        nome = str(nome).strip()
        if nome.upper() in {"TOTAIS", "TOTAL", ""}:
            continue
        flag = ws.cell(r, 5).value
        pays = {}
        for m, col in zip(MESES, COLS):
            v = ws.cell(r, openpyxl.utils.column_index_from_string(col)).value
            if v is not None:
                pays[m] = str(v).strip()
        hon = parse_honorario(ws.cell(r, 8).value, ws.cell(r, 7).value)
        ativo, fim = detectar(pays, hon)
        by_norm[norm(nome)] = {
            "nome": nome,
            "ativo": ativo,
            "fim": fim,
            "tipo": map_tipo(flag),
            "pays": pays,
            "hon": hon,
        }

    lines = [
        "BEGIN;",
        "ALTER TABLE cliente ADD COLUMN IF NOT EXISTS data_fim_cobranca DATE;",
    ]
    unmatched = []
    for cid, db_nome in DB_CLIENTES:
        info = by_norm.get(norm(db_nome))
        if not info:
            unmatched.append(db_nome)
            continue
        ativo = "TRUE" if info["ativo"] else "FALSE"
        fim = f"'{info['fim']}'" if info["fim"] and not info["ativo"] else "NULL"
        lines.append(
            f"UPDATE cliente SET ativo = {ativo}, tipo_pagamento = '{info['tipo']}', "
            f"data_fim_cobranca = {fim} WHERE id = {cid}; "
            f"-- {db_nome} <- {info['nome']}"
        )
        print(
            f"id={cid:2} {db_nome[:35]:35} -> "
            f"{'ATIVO' if info['ativo'] else 'INATIVO':7} tipo={info['tipo']:16} fim={info['fim']}"
        )

    lines.append("COMMIT;")
    out = Path(r"C:\Workspace\someli-contabilidade\deploy\import\fix_ativo_ids.sql")
    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nSQL: {out}")
    if unmatched:
        print("NÃO CASADOS:", unmatched)


if __name__ == "__main__":
    main()
