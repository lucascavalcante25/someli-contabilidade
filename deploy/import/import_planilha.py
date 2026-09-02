#!/usr/bin/env python3
"""
Importa dados reais da planilha FINANCEIRO SOMELI.xlsx para o Postgres do Someli.

Uso na VPS:
  pip install openpyxl psycopg2-binary
  python3 deploy/import/import_planilha.py \\
    --xlsx /root/FINANCEIRO_SOMELI.xlsx \\
    --host localhost --port 5432 --db someli --user someli --password '...'

Ou via Docker (recomendado):
  python3 deploy/import/import_planilha.py \\
    --xlsx /root/FINANCEIRO_SOMELI.xlsx \\
    --docker-container someli-contabo-db-1
"""

from __future__ import annotations

import argparse
import re
import subprocess
import sys
from decimal import Decimal
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

MESES_2026 = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}
COLS_MESES = "IJKLMNOPQRST"
SKIP_NOMES = {"TOTAIS", "TOTAL", ""}
ANO = 2026


def sql_literal(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    s = str(value).replace("'", "''")
    return f"'{s}'"


def map_tipo(flag) -> str:
    f = (flag or "S").strip().upper()
    if f == "P":
        return "pessoa_fisica"
    if f == "K":
        return "terceiros"
    return "pessoa_juridica"


def map_forma(forma) -> str | None:
    if not forma:
        return None
    f = str(forma).strip().upper()
    if f in {"-", "NONE"}:
        return None
    if "PIX" in f:
        return "pix"
    if "BOLETO" in f:
        return "boleto"
    return f.lower()[:30]


def parse_vencimento(v) -> int:
    if v is None:
        return 10
    if isinstance(v, (int, float)):
        d = int(v)
        return d if 1 <= d <= 31 else 10
    s = str(v).strip()
    if s.isdigit():
        d = int(s)
        return d if 1 <= d <= 31 else 10
    return 10


def parse_honorario(row_vals) -> Decimal:
    hon_h, hon_g = row_vals
    raw = hon_h if hon_h not in (None, "") else hon_g
    if raw is None or raw == "":
        return Decimal("0")
    try:
        return Decimal(str(raw).replace(",", "."))
    except Exception:
        return Decimal("0")


def is_ativo(flag, honorario, pays: dict) -> bool:
    f = (flag or "").strip().upper()
    if f == "P":
        return False
    ok = sum(1 for v in pays.values() if str(v).upper() == "OK")
    if honorario == 0 and ok == 0:
        return False
    return True


def parse_receitas(ws) -> list[dict]:
    clients = []
    last_resp = None
    last_ind = None
    seq = 0
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        nome = str(nome).strip()
        if nome.upper() in SKIP_NOMES:
            continue

        ind = ws.cell(r, 1).value
        resp = ws.cell(r, 2).value
        if ind:
            last_ind = str(ind).strip()
        if resp:
            last_resp = str(resp).strip()

        forma = ws.cell(r, 4).value
        flag = ws.cell(r, 5).value
        venc = ws.cell(r, 6).value
        honorario = parse_honorario((ws.cell(r, 7).value, ws.cell(r, 8).value))

        pays = {}
        for m, col in zip(MESES_2026, COLS_MESES):
            v = ws.cell(r, openpyxl.utils.column_index_from_string(col)).value
            if v is not None:
                pays[m] = str(v).strip()

        seq += 1
        cnpj = f"99{seq:012d}"
        clients.append({
            "cnpj": cnpj,
            "razao_social": nome,
            "nome_fantasia": nome,
            "honorario": honorario,
            "dia_vencimento": parse_vencimento(venc),
            "tipo_pagamento": map_tipo(flag),
            "forma_pagamento": map_forma(forma),
            "indicacao": last_ind,
            "responsavel_codigo": last_resp,
            "ativo": is_ativo(flag, honorario, pays),
            "pagamentos": pays,
        })
    return clients


def map_despesa_tipo(categoria) -> str:
    c = (categoria or "").strip().lower()
    if "cart" in c:
        return "cartao"
    if "pessoal" in c:
        return "recorrente"
    if "fixa" in c:
        return "fixa"
    return "recorrente"


def parse_despesas(ws) -> list[dict]:
    # header row 3, data from 4
    month_headers: list[tuple[int, int]] = []
    for c in range(4, ws.max_column + 1):
        h = ws.cell(3, c).value
        if not h:
            continue
        h_up = str(h).strip().upper()
        mes_map = {
            "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3, "ABRIL": 4,
            "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9,
            "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
        }
        for k, m in mes_map.items():
            if k in h_up.replace(" ", ""):
                month_headers.append((c, m))
                break

    despesas = []
    last_cat = None
    for r in range(4, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        dia = ws.cell(r, 2).value
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        nome = str(nome).strip()
        if nome.upper() in SKIP_NOMES:
            continue
        if cat:
            last_cat = str(cat).strip()

        valores = {}
        for col, mes in month_headers:
            v = ws.cell(r, col).value
            if v is None:
                continue
            try:
                val = Decimal(str(v).replace(",", "."))
            except Exception:
                continue
            if val > 0:
                valores[mes] = val

        if not valores:
            continue

        # valor mensal = último mês com valor ou média
        ultimo_valor = valores[max(valores)]
        despesas.append({
            "descricao": nome,
            "categoria": last_cat,
            "tipo": map_despesa_tipo(last_cat),
            "dia_pagamento": parse_vencimento(dia),
            "valor_mensal": ultimo_valor,
            "mensais": valores,
        })
    return despesas


def build_sql(clients: list[dict], despesas: list[dict]) -> str:
    lines = ["BEGIN;"]

    lines.append("DELETE FROM notification;")
    lines.append("DELETE FROM cliente_documento;")
    lines.append("DELETE FROM cliente_obrigacao;")
    lines.append("DELETE FROM pagamento_mensal;")
    lines.append("DELETE FROM despesa_mensal;")
    lines.append("DELETE FROM despesa;")
    lines.append("DELETE FROM cliente;")

    # Hemerson deve existir via migração V13 (Flyway na subida da API).

    for c in clients:
        resp_sql = "(SELECT id FROM usuario WHERE cpf = '22222222222' LIMIT 1)"

        lines.append(f"""
INSERT INTO cliente (
  cnpj, razao_social, nome_fantasia, honorario, dia_vencimento, tipo_pagamento,
  status, data_criacao, indicacao, forma_pagamento, ativo, responsavel_id
) VALUES (
  {sql_literal(c['cnpj'])},
  {sql_literal(c['razao_social'])},
  {sql_literal(c['nome_fantasia'])},
  {c['honorario']},
  {c['dia_vencimento']},
  {sql_literal(c['tipo_pagamento'])},
  'em_dia',
  NOW(),
  {sql_literal(c['indicacao'])},
  {sql_literal(c['forma_pagamento'])},
  {sql_literal(c['ativo'])},
  {resp_sql}
);
""")

        for mes_abbr, pago in c["pagamentos"].items():
            if str(pago).upper() != "OK":
                continue
            mes = MESES_2026[mes_abbr]
            lines.append(f"""
INSERT INTO pagamento_mensal (cliente_id, mes, ano, pago)
VALUES (
  (SELECT id FROM cliente WHERE cnpj = {sql_literal(c['cnpj'])}),
  {mes}, {ANO}, true
)
ON CONFLICT (cliente_id, mes, ano) DO UPDATE SET pago = true;
""")

    for d in despesas:
        desc = d["descricao"]
        lines.append(f"""
INSERT INTO despesa (descricao, valor_mensal, tipo, dia_pagamento, data_inicio, ativo, data_criacao, data_inicio_cobranca)
VALUES (
  {sql_literal(desc)},
  {d['valor_mensal']},
  {sql_literal(d['tipo'])},
  {d['dia_pagamento']},
  '2026-01-01',
  true,
  NOW(),
  '2026-01-01'
);
""")
        for mes, valor in d["mensais"].items():
            lines.append(f"""
INSERT INTO despesa_mensal (despesa_id, mes, ano, paga)
VALUES (
  (SELECT id FROM despesa WHERE descricao = {sql_literal(desc)} ORDER BY id DESC LIMIT 1),
  {mes}, {ANO}, true
)
ON CONFLICT (despesa_id, mes, ano) DO UPDATE SET paga = true;
""")

    lines.append("COMMIT;")
    return "\n".join(lines)


def run_psql_docker(container: str, sql: str, user: str, db: str) -> None:
    proc = subprocess.run(
        ["docker", "exec", "-i", container, "psql", "-U", user, "-d", db, "-v", "ON_ERROR_STOP=1"],
        input=sql.encode("utf-8"),
        capture_output=True,
    )
    if proc.returncode != 0:
        print(proc.stdout.decode())
        print(proc.stderr.decode(), file=sys.stderr)
        sys.exit(proc.returncode)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Caminho da planilha FINANCEIRO SOMELI.xlsx")
    parser.add_argument("--docker-container", default="someli-contabo-db-1")
    parser.add_argument("--user", default="someli")
    parser.add_argument("--db", default="someli")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"Arquivo não encontrado: {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    clients = parse_receitas(wb["RECEITAS 2026"])
    despesas = parse_despesas(wb["DESPESAS 2026"])

    print(f"Clientes: {len(clients)} | Despesas: {len(despesas)}")
    ativos = sum(1 for c in clients if c["ativo"])
    print(f"Clientes ativos: {ativos} | Inativos: {len(clients) - ativos}")

    sql = build_sql(clients, despesas)
    if args.dry_run:
        out = Path("import_planilha.sql")
        out.write_text(sql, encoding="utf-8")
        print(f"SQL gerado em {out.resolve()}")
        return

    run_psql_docker(args.docker_container, sql, args.user, args.db)
    print("Importação concluída com sucesso.")


if __name__ == "__main__":
    main()
