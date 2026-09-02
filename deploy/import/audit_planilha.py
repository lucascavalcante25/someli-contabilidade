import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r"C:\Users\Lucas Cavalcante\Downloads\FINANCEIRO SOMELI.xlsx", data_only=True)
ws = wb["RECEITAS 2026"]
MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
COLS = "IJKLMNOPQRST"


def honorario(r):
    h, g = ws.cell(r, 8).value, ws.cell(r, 7).value
    raw = h if h not in (None, "") else g
    try:
        return float(str(raw).replace(",", ".")) if raw not in (None, "") else 0.0
    except Exception:
        return 0.0


rows = []
for r in range(2, ws.max_row + 1):
    nome = ws.cell(r, 3).value
    if not nome:
        continue
    nome = str(nome).strip()
    if nome.upper() in {"TOTAIS", "TOTAL", ""}:
        continue
    flag = ws.cell(r, 5).value
    pays = {}
    for m, col in zip(MESES, COLS):
        v = ws.cell(r, openpyxl.utils.column_index_from_string(col)).value
        if v is not None:
            pays[m] = str(v).strip()
    hon = honorario(r)
    last_ok = None
    for m in MESES:
        if m in pays and pays[m].upper() == "OK":
            last_ok = m
    rows.append({"nome": nome, "flag": flag, "hon": hon, "pays": pays, "last_ok": last_ok})

print("flags", Counter(str(r["flag"]) for r in rows))
vals = Counter()
for r in rows:
    for v in r["pays"].values():
        vals[v] += 1
print("pay values", vals)
print()
for r in rows:
    ok_count = sum(1 for v in r["pays"].values() if v.upper() == "OK")
    print(
        f"{r['nome'][:42]:42} flag={str(r['flag']):3} hon={r['hon']:8.2f} ok={ok_count:2} last_ok={r['last_ok']} pays={r['pays']}"
    )
